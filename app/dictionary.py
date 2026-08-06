"""Chargement du dictionnaire : CSV multi-définitions (format studio) ou
Excel historique (une définition par mot)."""

from __future__ import annotations

import csv
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from app.models import Definition, Entry


def _normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().lower())
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def normalize_word(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().upper())
    return "".join(
        ch for ch in text if unicodedata.category(ch) != "Mn" and ch.isalpha()
    )


_ALIASES = {
    "word": {"mot", "reponse", "solution", "word", "answer"},
    "definition": {"definition", "indice", "question", "clue", "hint"},
    "display": {"affichage", "display", "forme"},
    "category": {"categorie", "category", "theme", "univers"},
    "register": {"registre", "register", "ton"},
    "level": {"niveau", "level", "difficulte"},
}


def _find_column(headers: list[str], names: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if header in names:
            return index
    return None


def load_dictionary(path: str | Path) -> list[Entry]:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        rows = _read_csv_rows(path)
    else:
        rows = _read_excel_rows(path)
    return _build_entries(rows)


def _read_csv_rows(path: Path):
    with open(path, encoding="utf-8-sig", newline="") as handle:
        yield from csv.reader(handle, delimiter=";")


def _read_excel_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        yield ["" if value is None else str(value) for value in row]


def _build_entries(rows) -> list[Entry]:
    rows = iter(rows)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise ValueError("Le fichier est vide.") from exc

    headers = [_normalize_header(value) for value in raw_headers]
    columns = {key: _find_column(headers, names) for key, names in _ALIASES.items()}

    if columns["word"] is None or columns["definition"] is None:
        raise ValueError(
            "Le fichier doit contenir une colonne Mot et une colonne Définition."
        )

    def cell(row: list, key: str) -> str:
        index = columns[key]
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    entries: dict[str, Entry] = {}

    for row in rows:
        word = normalize_word(cell(row, "word"))
        definition = cell(row, "definition")
        if len(word) < 2 or not definition:
            continue

        entry = entries.get(word)
        if entry is None:
            entry = Entry(
                word=word,
                display=cell(row, "display") or cell(row, "word"),
                category=cell(row, "category"),
            )
            entries[word] = entry

        if any(d.text == definition for d in entry.definitions):
            continue

        try:
            level = int(cell(row, "level") or 1)
        except ValueError:
            level = 1
        register = cell(row, "register") or "factuel"
        entry.definitions.append(
            Definition(text=definition, register=register, level=level)
        )

    result = [entry for entry in entries.values() if entry.definitions]
    if not result:
        raise ValueError("Aucune entrée exploitable dans le fichier.")
    return result
