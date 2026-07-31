from pathlib import Path
import unicodedata

from openpyxl import load_workbook

from app.models import WordEntry


def _normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().lower())
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def _normalize_word(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().upper())
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn" and ch.isalpha())


def load_dictionary(path: str | Path) -> list[WordEntry]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise ValueError("Le fichier Excel est vide.") from exc

    headers = [_normalize_header(value) for value in raw_headers]
    aliases = {
        "word": {"mot", "reponse", "solution", "word", "answer"},
        "definition": {"definition", "indice", "question", "clue", "hint"},
        "theme": {"theme", "categorie", "category", "univers"},
    }

    def find_column(names: set[str]) -> int | None:
        for index, header in enumerate(headers):
            if header in names:
                return index
        return None

    word_index = find_column(aliases["word"])
    definition_index = find_column(aliases["definition"])
    theme_index = find_column(aliases["theme"])

    if word_index is None or definition_index is None:
        raise ValueError("Le fichier doit contenir une colonne Mot et une colonne Définition.")

    entries: list[WordEntry] = []
    seen: set[str] = set()

    for row in rows:
        raw_word = row[word_index] if word_index < len(row) else None
        raw_definition = row[definition_index] if definition_index < len(row) else None
        raw_theme = row[theme_index] if theme_index is not None and theme_index < len(row) else ""

        word = _normalize_word(raw_word)
        definition = str(raw_definition or "").strip()
        theme = str(raw_theme or "").strip()

        if len(word) < 2 or not definition or word in seen:
            continue

        seen.add(word)
        entries.append(WordEntry(word=word, definition=definition, theme=theme))

    return entries
